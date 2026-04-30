"""One-shot script: build sample_tc_with_issues.xlsx for qa-review-tc tests.

Two tabs: TabA (8 rows with intentional intra-tab issues) + TabB (5 rows with
one cross-tab dup vs TabA).

Run once when setting up Phase 3.
"""
from pathlib import Path

from openpyxl import Workbook

TARGET = Path(__file__).parent.parent / "tests" / "fixtures" / "sample_tc_with_issues.xlsx"
HEADER = [
    "Priority", "OS", "Automation Check", "Test Item", "Automation TC_ID", "TC_ID",
    "Test Summary", "Remote Config / Admin", "Pre-condition", "Test Step",
    "Expected Result", "Result", "Jira no.", "Comment",
]


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # ---- TabA ----
    a = wb.create_sheet("TabA")
    a.append(HEADER)
    # Section header
    a.append([1.0, "메인 화면", "", "", "", "", "", "", "", "", "", "", "", ""])
    # Valid row
    a.append(["P1", "All", "All", "메인", "", "1-1", "메인 진입",
              "", "신규 사용자", "1. 앱 실행", "메인 화면 노출", "", "", ""])
    # Intra-tab dup of Test Summary
    a.append(["P2", "All", "Skip", "메인", "", "1-2", "메인 진입",
              "", "신규 사용자", "1. 앱 실행 후 진입", "메인 화면 노출", "", "", ""])
    # TC_ID DUPLICATE (intentional — 1-2 repeated)
    a.append(["P3", "All", "Skip", "메인", "", "1-2", "추천 카드 노출",
              "", "신규", "1. 라운지", "추천 노출", "", "", ""])
    # Missing Priority (포맷 위반)
    a.append(["", "All", "Skip", "메인", "", "1-3", "에러 처리",
              "", "오프라인", "1. 진입", "에러 노출", "", "", ""])
    # Empty Expected Result (포맷 위반)
    a.append(["P2", "All", "Skip", "메인", "", "1-4", "딥링크 진입",
              "", "딥링크 클릭", "1. 클릭", "", "", "", ""])
    # OS enum 위반
    a.append(["P3", "MacOS", "Skip", "메인", "", "1-5", "macOS 진입",
              "", "", "1. 실행", "노출", "", "", ""])
    # Cross-tab dup target — same Test Summary as TabB row
    a.append(["P2", "All", "Skip", "정책", "", "1-6", "차단 사용자 제외",
              "", "차단 상태", "1. 라운지 진입", "차단 사용자 비노출", "", "", ""])

    # ---- TabB ----
    b = wb.create_sheet("TabB")
    b.append(HEADER)
    b.append([1.0, "라운지", "", "", "", "", "", "", "", "", "", "", "", ""])
    b.append(["P1", "All", "All", "라운지", "", "1-1", "라운지 진입",
              "", "신규", "1. 라운지", "라운지 노출", "", "", ""])
    b.append(["P3", "All", "Skip", "라운지", "", "1-2", "스크롤 동작",
              "", "카드 2개+", "1. 스와이프", "스크롤 정상", "", "", ""])
    # Cross-tab dup with TabA 1-6
    b.append(["P2", "All", "Skip", "정책", "", "1-3", "차단 사용자 제외",
              "", "차단 상태", "1. 라운지 진입", "차단 사용자 비노출", "", "", ""])
    b.append(["P4", "All", "Skip", "에러", "", "1-4", "네트워크 실패",
              "", "오프라인", "1. 진입", "에러 처리", "", "", ""])

    TARGET.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TARGET)
    print(f"Wrote {TARGET}")


if __name__ == "__main__":
    main()
