"""Tests for shared/summary_xlsx.py."""
from pathlib import Path

import pytest
from openpyxl import load_workbook

from shared.summary_xlsx import write_summary_workbook

SHEETS = [
    {"title": "리스크 매트릭스",
     "headers": ["영역", "등급"],
     "rows": [["결제", "Blocker"], ["UI 카피", "Minor"]]},
    {"title": "테스트 포커스",
     "headers": ["영역", "권장 TC"],
     "rows": [["결제", "구매 실패 복구"]]},
]


def test_writes_sheets_in_order(tmp_path: Path):
    out = write_summary_workbook(SHEETS, tmp_path / "summary.xlsx")
    wb = load_workbook(out)
    assert wb.sheetnames == ["리스크 매트릭스", "테스트 포커스"]


def test_writes_headers_and_rows(tmp_path: Path):
    out = write_summary_workbook(SHEETS, tmp_path / "summary.xlsx")
    ws = load_workbook(out)["리스크 매트릭스"]
    assert [c.value for c in ws[1]] == ["영역", "등급"]
    assert [c.value for c in ws[2]] == ["결제", "Blocker"]
    assert [c.value for c in ws[3]] == ["UI 카피", "Minor"]


def test_collision_appends_suffix(tmp_path: Path):
    first = write_summary_workbook(SHEETS, tmp_path / "summary.xlsx")
    second = write_summary_workbook(SHEETS, tmp_path / "summary.xlsx")
    assert first.name == "summary.xlsx"
    assert second.name == "summary (2).xlsx"


def test_empty_sheets_raises(tmp_path: Path):
    with pytest.raises(ValueError):
        write_summary_workbook([], tmp_path / "summary.xlsx")
