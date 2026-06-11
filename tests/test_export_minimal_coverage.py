"""Tests for shared/export_minimal_coverage.py."""
from pathlib import Path

from openpyxl import load_workbook

from shared.export_minimal_coverage import ANALYSIS_COLUMNS, export_minimal_coverage
from shared.select_minimal_coverage import select_minimal_coverage

EXPECTED_SHEETS = ["Selected TC", "Coverage Summary", "Excluded TC",
                   "Next Best", "Assumptions"]


def _mutual_row(tc_id, summary, priority="P2", os="", item="매치"):
    """Mutual 탭 형태: 14컬럼 변형 + A/B 컬럼, Test Step → Test Reproduce.

    item 기본값 "매치"는 고위험 키워드라 강제 포함됨 — 비강제 행이 필요한
    테스트는 비위험 값(예: "설정")을 넘길 것.
    """
    return {
        "Priority": priority, "OS": os, "Test Item": item, "TC_ID": tc_id,
        "Test Summary": summary, "Remote Config / Admin": "",
        "Pre-condition": "", "A": "발신", "B": "수신",
        "Test Reproduce": "1. A가 매치 시작", "Expected Result": "ok",
        "Result": "", "Jira no.": "", "Comment": "",
        "사내메모": "내부 전용",  # 비표준 컬럼 — 범용 보존 검증용
    }


def _selection(rows):
    return select_minimal_coverage(rows)


COLUMNS = ["Priority", "OS", "Test Item", "TC_ID", "Test Summary",
           "Remote Config / Admin", "Pre-condition", "A", "B",
           "Test Reproduce", "Expected Result", "Result", "Jira no.", "Comment",
           "사내메모"]


def test_writes_five_sheets(tmp_path: Path):
    rows = [_mutual_row("1-1", "매치 성사"), _mutual_row("1-2", "매치 거절")]
    out = export_minimal_coverage(_selection(rows), COLUMNS, tmp_path / "min.xlsx")
    assert load_workbook(out).sheetnames == EXPECTED_SHEETS


def test_preserves_all_source_columns_and_tc_id(tmp_path: Path):
    rows = [_mutual_row("1-1", "매치 성사")]
    out = export_minimal_coverage(_selection(rows), COLUMNS, tmp_path / "min.xlsx")
    ws = load_workbook(out)["Selected TC"]
    headers = [c.value for c in ws[1]]
    assert headers == COLUMNS + list(ANALYSIS_COLUMNS)   # mutual A/B 포함 범용 보존
    row2 = {h: c.value for h, c in zip(headers, ws[2])}
    assert row2["TC_ID"] == "1-1"                        # 원본 TC_ID 유지
    assert row2["실행 순서"] == 1


def test_excluded_sheet_has_reason(tmp_path: Path):
    rows = [_mutual_row("1-1", "매치 성사")] + [
        _mutual_row(f"2-{i}", "라벨 확인", priority="P4", item="설정") for i in range(8)]
    selection = select_minimal_coverage(rows, next_best_count=1)
    out = export_minimal_coverage(selection, COLUMNS, tmp_path / "min.xlsx")
    ws = load_workbook(out)["Excluded TC"]
    assert ws.max_row >= 2
    # 헤더: ["TC_ID", "Test Summary", "제외 사유", "잔여 리스크", "강제 대상"]
    headers = [c.value for c in ws[1]]
    assert headers == ["TC_ID", "Test Summary", "제외 사유", "잔여 리스크", "강제 대상"]
    assert ws.cell(row=2, column=3).value  # 제외 사유 비어있지 않음
    # forced-overflow가 아닌 일반 제외 행은 강제 대상 컬럼이 "" 또는 빈 값
    for row_idx in range(2, ws.max_row + 1):
        forced_val = ws.cell(row=row_idx, column=5).value
        assert forced_val in (None, "", False, "Y"), f"row {row_idx}: unexpected forced={forced_val!r}"
    # forced_overflow가 없는 케이스라 모두 "" 이어야 함
    for row_idx in range(2, ws.max_row + 1):
        assert ws.cell(row=row_idx, column=5).value in (None, "")
