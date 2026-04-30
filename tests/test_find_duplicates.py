"""Tests for shared/find_duplicates.py."""
import json
import subprocess
import sys
from pathlib import Path

import pytest


@pytest.fixture
def issues_xlsx_path(fixtures_dir: Path) -> Path:
    return fixtures_dir / "sample_tc_with_issues.xlsx"


def _run(xlsx: Path, tab: str, no_cross: bool = False) -> dict:
    args = [sys.executable, "shared/find_duplicates.py", str(xlsx), "--tab", tab]
    if no_cross:
        args.append("--no-cross-tab")
    result = subprocess.run(args, capture_output=True, text=True, check=True)
    return json.loads(result.stdout)


class TestFindDuplicates:
    def test_returns_intra_and_cross_keys(self, issues_xlsx_path: Path):
        out = _run(issues_xlsx_path, "TabA")
        assert "intra_tab" in out
        assert "cross_tab" in out

    def test_detects_intra_tab_test_summary_dup(self, issues_xlsx_path: Path):
        out = _run(issues_xlsx_path, "TabA")
        # TabA has '메인 진입' on rows 3 and 4 (1-1 and 1-2)
        intra = [d for d in out["intra_tab"] if d["field"] == "Test Summary"
                 and "메인 진입" in d.get("value", "")]
        assert len(intra) >= 1

    def test_detects_cross_tab_test_summary_dup(self, issues_xlsx_path: Path):
        out = _run(issues_xlsx_path, "TabA")
        # TabA 1-6 ↔ TabB 1-3 share '차단 사용자 제외'
        cross = [d for d in out["cross_tab"]
                 if "차단 사용자" in d.get("value", "")]
        assert len(cross) >= 1
        assert all(d["focus_tab"] == "TabA" and d["other_tab"] == "TabB"
                   for d in cross)

    def test_no_cross_flag_disables_cross_scan(self, issues_xlsx_path: Path):
        out = _run(issues_xlsx_path, "TabA", no_cross=True)
        assert out["cross_tab"] == []

    def test_summary_tabs_excluded_from_cross_scan(self, tmp_path: Path):
        # Build a synthetic xlsx with TabA + Summary
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
        for tab_name in ("TabA", "Summary"):
            ws = wb.create_sheet(tab_name)
            ws.append(["Priority", "OS", "Automation Check", "Test Item",
                       "Automation TC_ID", "TC_ID", "Test Summary",
                       "Remote Config / Admin", "Pre-condition", "Test Step",
                       "Expected Result", "Result", "Jira no.", "Comment"])
            ws.append([1.0, "x", "", "", "", "", "", "", "", "", "", "", "", ""])
            ws.append(["P1", "All", "All", "x", "", "1-1", "공통 텍스트",
                       "", "x", "x", "x", "", "", ""])
        path = tmp_path / "sm.xlsx"
        wb.save(path)
        out = _run(path, "TabA")
        # Summary tab's row should NOT appear in cross_tab
        summary_hits = [d for d in out["cross_tab"] if d["other_tab"] == "Summary"]
        assert summary_hits == []

    def test_cross_tab_skips_short_test_step_noise(self, tmp_path: Path):
        """Boilerplate Test Steps like '1. 진입' must NOT trigger cross-tab dup
        flags. Otherwise a 28-tab master floods the report with false positives."""
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
        header = ["Priority", "OS", "Automation Check", "Test Item",
                  "Automation TC_ID", "TC_ID", "Test Summary",
                  "Remote Config / Admin", "Pre-condition", "Test Step",
                  "Expected Result", "Result", "Jira no.", "Comment"]
        for tab_name in ("TabX", "TabY"):
            ws = wb.create_sheet(tab_name)
            ws.append(header)
            ws.append([1.0, "x", "", "", "", "", "", "", "", "", "", "", "", ""])
            ws.append(["P1", "All", "All", "x", "", "1-1", f"{tab_name} 고유 시나리오",
                       "", "x", "1. 진입", "x", "", "", ""])
        path = tmp_path / "noise.xlsx"
        wb.save(path)
        out = _run(path, "TabX")
        # Same boilerplate Test Step "1. 진입" in both tabs — must NOT flag.
        step_hits = [d for d in out["cross_tab"]
                     if d["field"] == "Test Step (normalized)"]
        assert step_hits == [], \
            f"Short boilerplate Test Step incorrectly flagged: {step_hits}"
