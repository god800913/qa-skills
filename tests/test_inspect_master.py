"""Tests for shared/inspect_master.py."""
from pathlib import Path

import pytest

from shared.inspect_master import list_tabs
from shared.inspect_master import parse_tab_meta


class TestListTabs:
    def test_returns_all_tabs_with_summary_marked_excluded(self, minimal_master_path: Path):
        tabs = list_tabs(minimal_master_path)
        names = {t["name"] for t in tabs}
        assert names == {"login", "Lounge"}
        for tab in tabs:
            assert "is_summary" in tab
            assert "column_count" in tab
            assert tab["is_summary"] is False  # neither is a Summary tab


class TestListTabsSummaryDetection:
    def test_summary_tab_marked_excluded(self, tmp_path: Path):
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Summary").append(["x"])
        wb.create_sheet("Summary의 사본").append(["x"])
        wb.create_sheet("Lounge").append(["Priority", "OS"])
        path = tmp_path / "synthetic.xlsx"
        wb.save(path)

        tabs = list_tabs(path)
        names_to_exclusion = {t["name"]: t["is_summary"] for t in tabs}
        assert names_to_exclusion == {
            "Summary": True,
            "Summary의 사본": True,
            "Lounge": False,
        }


class TestParseTabMeta:
    def test_login_tab_has_leading_blank_column_handled(self, minimal_master_path: Path):
        meta = parse_tab_meta(minimal_master_path, "login")
        # Standard columns must be present in mapping regardless of leading blank
        assert "Priority" in meta["columns"]
        assert "TC_ID" in meta["columns"]
        # The mapping value is the actual cell index in the row, so leading-blank tabs
        # have Priority at index 1, not 0.
        assert meta["columns"]["Priority"] >= 1

    def test_lounge_tab_no_leading_blank(self, minimal_master_path: Path):
        meta = parse_tab_meta(minimal_master_path, "Lounge")
        assert "Priority" in meta["columns"]
        # Lounge should have Priority at index 0 (no leading blank)
        assert meta["columns"]["Priority"] == 0

    def test_returns_template_type_single(self, minimal_master_path: Path):
        meta = parse_tab_meta(minimal_master_path, "Lounge")
        assert meta["template_type"] == "single"


def make_mutual_xlsx(tmp_path: Path) -> Path:
    """Build a mutual-template tab (Test Reproduce + A/B columns) like real `in Match` tabs."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("in Match")
    ws.append([
        "Priority", "OS", "Automation Check", "Test Item", "Automation TC_ID",
        "TC_ID", "Test Summary", "Remote Config / Admin", "Pre-condition",
        "Test Reproduce", "Expected Result", "A", "B", "Result", "Jira no.", "Comment",
    ])
    ws.append([1.0, "", "", "매치 기본", "", "", "", "", "", "", "", "", "", "", "", ""])
    ws.append([
        "P1", "All", "Skip", "매치 진입", "", "1-1", "양측 매치 연결", "",
        "두 계정 로그인", "A: 매치 시작\nB: 매치 수락", "영상 연결됨",
        "기기A", "기기B", "", "", "",
    ])
    path = tmp_path / "mutual.xlsx"
    wb.save(path)
    return path


class TestParseTabMetaMutual:
    def test_template_type_detected_as_mutual(self, tmp_path: Path):
        meta = parse_tab_meta(make_mutual_xlsx(tmp_path), "in Match")
        assert meta["template_type"] == "mutual"

    def test_test_reproduce_maps_to_canonical_test_step(self, tmp_path: Path):
        meta = parse_tab_meta(make_mutual_xlsx(tmp_path), "in Match")
        # "Test Reproduce" must canonicalize to "Test Step" so downstream
        # consumers (validate_format, find_duplicates, select_minimal_coverage)
        # see a single canonical key.
        assert "Test Step" in meta["columns"]
        assert "Test Reproduce" not in meta["columns"]
        assert meta["columns"]["Test Step"] == 9


class TestParseTabMetaErrors:
    def test_header_not_found_error_includes_tab_and_file_context(self, tmp_path: Path):
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("BrokenTab")
        ws.append(["foo", "bar"])  # no Priority cell
        path = tmp_path / "broken.xlsx"
        wb.save(path)

        with pytest.raises(ValueError, match="BrokenTab"):
            parse_tab_meta(path, "BrokenTab")


class TestParseTabMetaSections:
    def test_lounge_has_sections(self, minimal_master_path: Path):
        meta = parse_tab_meta(minimal_master_path, "Lounge")
        # Lounge fixture has at least one section header in the first 30 rows
        assert "sections" in meta
        assert isinstance(meta["sections"], list)
        # Each section has name, header_row, last_tc_id (or None)
        for sec in meta["sections"]:
            assert "name" in sec
            assert "header_row" in sec
            assert "last_tc_id" in sec  # may be None if no TCs in section yet

    def test_last_tc_id_format(self, minimal_master_path: Path):
        meta = parse_tab_meta(minimal_master_path, "Lounge")
        ids_found = [s["last_tc_id"] for s in meta["sections"] if s["last_tc_id"]]
        # Pattern <section>-<number>, e.g. "1-12"
        for tc_id in ids_found:
            assert "-" in tc_id, f"Unexpected TC_ID shape: {tc_id}"

    def test_login_section_name_skips_dash_sentinel(self, minimal_master_path: Path):
        meta = parse_tab_meta(minimal_master_path, "login")
        # The login tab has '-' as Priority-1 cell in the section header row;
        # the real section label is in a later cell. Names should not be just '-'.
        names = [s["name"] for s in meta["sections"]]
        for n in names:
            assert n != "-", f"Section name '-' is the dash sentinel, not a real label"
            assert n.strip(), "Section name should not be empty"


import json
import subprocess
import sys


class TestCLI:
    def test_no_tab_lists_all(self, minimal_master_path: Path):
        result = subprocess.run(
            [sys.executable, "shared/inspect_master.py", str(minimal_master_path)],
            capture_output=True,
            text=True,
            check=True,
        )
        data = json.loads(result.stdout)
        assert "tabs" in data
        names = {t["name"] for t in data["tabs"]}
        assert names == {"login", "Lounge"}

    def test_with_tab_returns_meta(self, minimal_master_path: Path):
        result = subprocess.run(
            [sys.executable, "shared/inspect_master.py", str(minimal_master_path), "--tab", "Lounge"],
            capture_output=True,
            text=True,
            check=True,
        )
        data = json.loads(result.stdout)
        assert data["tab"] == "Lounge"
        assert "columns" in data
        assert "sections" in data


class TestParseTabMetaSampleRows:
    def test_returns_sample_rows(self, minimal_master_path: Path):
        meta = parse_tab_meta(minimal_master_path, "Lounge")
        assert "sample_rows" in meta
        assert isinstance(meta["sample_rows"], list)
        # By default 3 rows
        assert 1 <= len(meta["sample_rows"]) <= 3

    def test_sample_rows_are_tc_rows_not_section_headers(self, minimal_master_path: Path):
        meta = parse_tab_meta(minimal_master_path, "Lounge")
        # Each sample row should have a TC_ID matching the pattern
        for row in meta["sample_rows"]:
            tc_id_idx = meta["columns"].get("TC_ID")
            assert tc_id_idx is not None
            tc_id = str(row[tc_id_idx]) if tc_id_idx < len(row) else ""
            assert "-" in tc_id, f"Sample row TC_ID looks wrong: {tc_id}"
