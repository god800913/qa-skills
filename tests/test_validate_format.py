"""Tests for shared/validate_format.py."""
import json
import subprocess
import sys
from pathlib import Path

import pytest


@pytest.fixture
def issues_xlsx_path(fixtures_dir: Path) -> Path:
    return fixtures_dir / "sample_tc_with_issues.xlsx"


def _run(xlsx: Path, tab: str) -> dict:
    result = subprocess.run(
        [sys.executable, "shared/validate_format.py", str(xlsx), "--tab", tab],
        capture_output=True, text=True, check=True,
    )
    return json.loads(result.stdout)


class TestValidateFormat:
    def test_returns_dict_with_issues_and_summary(self, issues_xlsx_path: Path):
        out = _run(issues_xlsx_path, "TabA")
        assert "issues" in out
        assert "summary" in out
        assert "total_rows" in out["summary"]

    def test_detects_missing_priority(self, issues_xlsx_path: Path):
        out = _run(issues_xlsx_path, "TabA")
        msgs = [i for i in out["issues"]
                if i["category"] == "missing_required" and i["field"] == "Priority"]
        assert len(msgs) == 1, f"Expected 1 missing Priority, got: {msgs}"

    def test_detects_missing_expected_result(self, issues_xlsx_path: Path):
        out = _run(issues_xlsx_path, "TabA")
        msgs = [i for i in out["issues"]
                if i["category"] == "missing_required" and i["field"] == "Expected Result"]
        assert len(msgs) == 1

    def test_detects_invalid_os_enum(self, issues_xlsx_path: Path):
        out = _run(issues_xlsx_path, "TabA")
        msgs = [i for i in out["issues"]
                if i["category"] == "invalid_enum" and i["field"] == "OS"]
        assert len(msgs) == 1
        assert "MacOS" in msgs[0]["message"]

    def test_detects_duplicate_tc_id(self, issues_xlsx_path: Path):
        out = _run(issues_xlsx_path, "TabA")
        msgs = [i for i in out["issues"] if i["category"] == "duplicate_tc_id"]
        assert len(msgs) >= 1
        assert any("1-2" in m["message"] for m in msgs)

    def test_each_issue_has_row_number(self, issues_xlsx_path: Path):
        out = _run(issues_xlsx_path, "TabA")
        for issue in out["issues"]:
            assert "row" in issue
            assert isinstance(issue["row"], int)
            assert issue["row"] > 0

    def test_string_form_section_header_skipped(self, tmp_path: Path):
        """Real masters use string section headers like '1. 라운지 메인' instead of
        numeric. Without unified _is_section_header, those would be treated as TC
        rows and generate false-positive missing_required issues."""
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("StringSec")
        ws.append(["Priority", "OS", "Automation Check", "Test Item",
                   "Automation TC_ID", "TC_ID", "Test Summary",
                   "Remote Config / Admin", "Pre-condition", "Test Step",
                   "Expected Result", "Result", "Jira no.", "Comment"])
        # String-form section header — must be skipped
        ws.append(["1. 라운지 메인", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        # Real TC row
        ws.append(["P1", "All", "All", "라운지", "", "1-1", "라운지 진입",
                   "", "신규", "1. 라운지", "라운지 노출", "", "", ""])
        path = tmp_path / "string_sec.xlsx"
        wb.save(path)

        out = _run(path, "StringSec")
        # If section header weren't skipped, we'd see missing_required issues for
        # the section row. Total TC rows must be exactly 1.
        assert out["summary"]["total_rows"] == 1, \
            f"Expected 1 TC row (section skipped), got {out['summary']}"
        # No missing_required issues — the one TC row is fully populated.
        missing = [i for i in out["issues"] if i["category"] == "missing_required"]
        assert missing == [], f"Unexpected missing_required: {missing}"


class TestAutomationCheckNotValidated:
    def test_arbitrary_automation_check_value_produces_no_issue(self, tmp_path: Path):
        """Automation Check는 사람-소관 컬럼 — validate_format이 enum 검사하지 않는다."""
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("AutoFree")
        ws.append(["Priority", "OS", "Automation Check", "Test Item",
                   "Automation TC_ID", "TC_ID", "Test Summary",
                   "Remote Config / Admin", "Pre-condition", "Test Step",
                   "Expected Result", "Result", "Jira no.", "Comment"])
        ws.append(["P1", "All", "임의값", "라운지", "", "1-1", "라운지 진입",
                   "", "신규", "1. 라운지", "라운지 노출", "", "", ""])
        path = tmp_path / "auto_free.xlsx"
        wb.save(path)

        out = _run(path, "AutoFree")
        auto_issues = [i for i in out["issues"] if i.get("field") == "Automation Check"]
        assert auto_issues == [], f"Automation Check 검사 잔존: {auto_issues}"


class TestValidateFormatMutual:
    def test_filled_test_reproduce_not_flagged_missing(self, tmp_path: Path):
        from tests.test_inspect_master import make_mutual_xlsx

        out = _run(make_mutual_xlsx(tmp_path), "in Match")
        step_issues = [i for i in out["issues"]
                       if i["category"] == "missing_required" and i["field"] == "Test Step"]
        assert step_issues == [], (
            "Test Reproduce column must satisfy the Test Step requirement on mutual tabs"
        )
