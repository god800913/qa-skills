"""Create a new TC workbook from a rows JSON.

Usage:
    python new_workbook.py --rows rows.json --output out.xlsx \
        --tab-name "Lounge 신규" [--template single|mutual]

rows.json schema:
    {"rows": [{"section": "1. 라운지 메인", "Priority": "P1", ...}, ...]}

Each row's "section" field groups it. TC_IDs are auto-assigned per section
starting at <section_idx>-1.

If output path already exists, appends "(2)", "(3)", ... suffix.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Standard column order for the output sheet (single template).
SINGLE_COLUMNS = (
    "Priority", "OS", "Automation Check", "Test Item",
    "Automation TC_ID", "TC_ID", "Test Summary",
    "Remote Config / Admin", "Pre-condition", "Test Step", "Expected Result",
    "Result", "Jira no.", "Comment",
)

MUTUAL_EXTRA_COLUMNS = ("A", "B")

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
SECTION_FONT = Font(bold=True)
SECTION_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")


def _resolve_output_path(path: Path) -> Path:
    """If path exists, append (2), (3), ... before suffix."""
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    n = 2
    while True:
        candidate = parent / f"{stem} ({n}){suffix}"
        if not candidate.exists():
            return candidate
        n += 1


def _section_index(name: str) -> int:
    """Extract leading numeric section index from name like '1. 라운지 메인' → 1.
    Returns sequential index (1, 2, 3, ...) based on first-seen order if not parseable."""
    m = re.match(r"^(\d+)", name)
    if m:
        return int(m.group(1))
    return 0  # caller handles fallback


def _columns_for(template: str) -> tuple[str, ...]:
    if template == "mutual":
        # Insert A/B before Result.
        # KNOWN GAP (Phase 2 PoC): spec §6.2 also requires renaming "Test Step" → "Test Reproduce"
        # for mutual templates. Not implemented here — mutual append on a real `in Match`-style tab
        # would currently write blank Test Reproduce cells because row_data uses "Test Step" key.
        # Fix in a follow-up: either add a TEMPLATE_COLUMN_RENAMES map or have callers normalize keys.
        idx = SINGLE_COLUMNS.index("Result")
        return SINGLE_COLUMNS[:idx] + MUTUAL_EXTRA_COLUMNS + SINGLE_COLUMNS[idx:]
    return SINGLE_COLUMNS


def write_workbook(rows: list[dict], tab_name: str, output: Path,
                   template: str = "single") -> Path:
    """Create a new xlsx with one tab. Returns the actual path written
    (may differ from `output` due to collision suffix)."""
    actual_output = _resolve_output_path(output)
    columns = _columns_for(template)

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(tab_name)

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP

    # Group rows by section, preserving first-seen order
    sections: dict[str, list[dict]] = {}
    for r in rows:
        s = r.get("section") or "(default)"
        sections.setdefault(s, []).append(r)

    excel_row = 2
    for section_idx, (section_name, section_rows) in enumerate(sections.items(), start=1):
        # Section header row — write the index in Priority column, name in Test Item col
        sec_pri_cell = ws.cell(row=excel_row, column=columns.index("Priority") + 1,
                               value=float(section_idx))
        sec_pri_cell.font = SECTION_FONT
        sec_pri_cell.fill = SECTION_FILL
        sec_name_cell = ws.cell(row=excel_row, column=columns.index("Test Item") + 1,
                                value=section_name)
        sec_name_cell.font = SECTION_FONT
        sec_name_cell.fill = SECTION_FILL
        excel_row += 1

        # TC rows
        for tc_seq, row_data in enumerate(section_rows, start=1):
            tc_id = f"{section_idx}-{tc_seq}"
            for col_idx, col_name in enumerate(columns, start=1):
                if col_name == "TC_ID":
                    value = tc_id
                else:
                    value = row_data.get(col_name, "")
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.alignment = WRAP
            excel_row += 1

    actual_output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(actual_output)
    return actual_output


def main() -> None:
    parser = argparse.ArgumentParser(description="Create a new TC workbook.")
    parser.add_argument("--rows", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--tab-name", type=str, default="Sheet1")
    parser.add_argument("--template", choices=["single", "mutual"], default="single")
    args = parser.parse_args()

    data = json.loads(args.rows.read_text(encoding="utf-8"))
    rows = data.get("rows", [])
    actual = write_workbook(rows, args.tab_name, args.output, template=args.template)
    print(f"Wrote {actual}")
    print(actual)


if __name__ == "__main__":
    main()
