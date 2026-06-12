# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1.0", "python-calamine>=0.2.0"]
# ///
"""Patch duplicate TC_IDs in an xlsx tab (qa:review-tc --patch).

Second and later occurrences of a duplicated TC_ID get a `-dup-N` suffix
(N = occurrence number, e.g. `1-2` → `1-2-dup-2`). The first occurrence
keeps its original ID. The input file is NEVER modified — output is always
a new file (collision-safe `(2)`, `(3)` suffix).

CLI:
    python patch_tc_ids.py <xlsx_path> --tab <tab_name> --output <out.xlsx>

stdout: patch summary, then the actual output path as the last line.
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

# Sibling import
sys.path.insert(0, str(Path(__file__).parent))
from inspect_master import parse_tab_meta  # noqa: E402


def _resolve_output_path(path: Path) -> Path:
    """If path exists, append (2), (3), ... before suffix.

    Duplicated from new_workbook.py on purpose: the qa-review-tc bundle
    ships without new_workbook.py, and this 10-line helper is not worth
    dragging the whole workbook writer in for.
    """
    if not path.exists():
        return path
    n = 2
    while True:
        candidate = path.parent / f"{path.stem} ({n}){path.suffix}"
        if not candidate.exists():
            return candidate
        n += 1


def patch_tc_ids(xlsx_path: Path, tab_name: str, output: Path) -> tuple[Path, list[dict]]:
    """Return (actual_output_path, patches) where each patch is
    {"row": excel_row, "old": str, "new": str}."""
    meta = parse_tab_meta(xlsx_path, tab_name)
    tc_id_col = meta["columns"].get("TC_ID")
    if tc_id_col is None:
        raise ValueError(f"Tab '{tab_name}' has no TC_ID column")
    header_row_idx = meta["header_row"]

    wb = load_workbook(xlsx_path)
    ws = wb[tab_name]

    seen: Counter[str] = Counter()
    patches: list[dict] = []
    for row_idx in range(header_row_idx + 2, ws.max_row + 1):  # openpyxl is 1-based
        cell = ws.cell(row=row_idx, column=tc_id_col + 1)
        value = cell.value
        if value is None or not str(value).strip():
            continue
        tc_id = str(value).strip()
        seen[tc_id] += 1
        if seen[tc_id] > 1:
            new_id = f"{tc_id}-dup-{seen[tc_id]}"
            cell.value = new_id
            patches.append({"row": row_idx, "old": tc_id, "new": new_id})

    actual_output = _resolve_output_path(output)
    actual_output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(actual_output)
    return actual_output, patches


def main() -> None:
    parser = argparse.ArgumentParser(description="Patch duplicate TC_IDs into a new xlsx.")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--tab", type=str, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    actual, patches = patch_tc_ids(args.xlsx_path, args.tab, args.output)
    print(f"Patched {len(patches)} duplicate TC_ID(s)")
    for p in patches:
        print(f"  row {p['row']}: {p['old']} -> {p['new']}")
    print(actual)


if __name__ == "__main__":
    main()
