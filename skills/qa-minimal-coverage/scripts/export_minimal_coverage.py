"""Export a minimal-coverage selection to a 5-sheet workbook.

Sheets: Selected TC / Coverage Summary / Excluded TC / Next Best / Assumptions.
Preserves ALL source columns generically (mutual tabs incl. A/B survive),
keeps original TC_ID values, appends analysis columns after source columns.
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from new_workbook import HEADER_FILL, HEADER_FONT, WRAP, _resolve_output_path  # noqa: E402
from select_minimal_coverage import risk_tags  # noqa: E402

from openpyxl import Workbook  # noqa: E402

ANALYSIS_COLUMNS = ("실행 순서", "선택 사유", "커버 리스크", "점수")


def export_minimal_coverage(selection: dict, columns: list[str], output: Path) -> Path:
    """Write the 5-sheet workbook. Returns actual path written."""
    actual_output = _resolve_output_path(output)
    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(title: str, headers: list[str], rows: list[list]) -> None:
        ws = wb.create_sheet(title)
        for col_idx, name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = WRAP
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value).alignment = WRAP

    add_sheet(
        "Selected TC",
        list(columns) + list(ANALYSIS_COLUMNS),
        [[item["row"].get(c) for c in columns]
         + [order, "; ".join(item["reasons"]),
            ", ".join(item["new_tags"]), item["score"]]
         for order, item in enumerate(selection["selected"], start=1)],
    )

    counts: dict[str, int] = {}
    for item in selection["selected"]:
        for tag in risk_tags(item["row"]):
            counts[tag] = counts.get(tag, 0) + 1
    add_sheet("Coverage Summary", ["리스크 태그", "커버 TC 수"],
              [[tag, n] for tag, n in sorted(counts.items())])

    add_sheet("Excluded TC", ["TC_ID", "Test Summary", "제외 사유", "잔여 리스크", "강제 대상"],
              [[e["row"].get("TC_ID"), e["row"].get("Test Summary"),
                e["reason"], ", ".join(e["residual_risk"]),
                "Y" if e.get("forced_overflow") else ""]
               for e in selection["excluded"]])

    add_sheet("Next Best", ["TC_ID", "Test Summary", "점수", "추가 커버"],
              [[n["row"].get("TC_ID"), n["row"].get("Test Summary"),
                n["score"], ", ".join(n["new_tags"])]
               for n in selection["next_best"]])

    add_sheet("Assumptions", ["가정"], [[a] for a in selection["assumptions"]])

    actual_output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(actual_output)
    return actual_output
