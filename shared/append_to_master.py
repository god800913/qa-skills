"""Append TC rows to a specific tab of an existing master xlsx.

Usage:
    python append_to_master.py --master master.xlsx --tab "Lounge" \
        --rows rows.json --output out.xlsx

NEVER modifies the master. Always writes a new file (collision-safe).
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Local imports — shared/inspect_master.py for column mapping + section info
sys.path.insert(0, str(Path(__file__).parent))
from inspect_master import parse_tab_meta  # noqa: E402

WRAP = Alignment(wrap_text=True, vertical="top")


def _find_section(user_section: str, sections: list[dict]) -> dict | None:
    """Find section matching user-supplied name.

    Match rules (unidirectional, by design):
    1. Exact match.
    2. Stored section name appears verbatim in user's name (user typed a more
       specific variant of an existing section).
    3. Otherwise None — caller should treat as new section.

    NOTE: This is intentionally NOT bidirectional. We do NOT match when user's
    name is a substring of stored name, because that causes false positives
    (e.g. "Lounge" wrongly matching "Lounge Navigation").
    """
    for s in sections:
        if s["name"] == user_section:
            return s
    for s in sections:
        if s["name"] in user_section:
            return s
    return None


def _resolve_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    n = 2
    while True:
        candidate = path.parent / f"{stem} ({n}){suffix}"
        if not candidate.exists():
            return candidate
        n += 1


def _next_tc_id(section_idx_str: str, last_tc_id: str | None) -> str:
    """Given e.g. '1' and last '1-23', return '1-24'. If no last, start at <idx>-1."""
    if last_tc_id:
        stripped = last_tc_id.strip()
        if "-" in stripped:
            sec, num = stripped.split("-", 1)
            if num.strip().isdigit():
                return f"{sec.strip()}-{int(num.strip()) + 1}"
    return f"{section_idx_str}-1"


def append_to_master(master: Path, tab: str, rows: list[dict], output: Path) -> Path:
    """Append rows to `tab` of `master`. Writes to `output` (or collision-safe variant).
    Returns actual output path."""
    meta = parse_tab_meta(master, tab)
    columns: dict[str, int] = meta["columns"]
    sections: list[dict] = meta["sections"]

    # Make output dir + collision-safe path
    actual_output = _resolve_output_path(output)
    actual_output.parent.mkdir(parents=True, exist_ok=True)

    # Open with openpyxl (preserves formatting, supports append)
    wb = load_workbook(master)
    if tab not in wb.sheetnames:
        raise ValueError(f"Tab '{tab}' not found in {master.name}. Available: {wb.sheetnames}")
    ws = wb[tab]

    # Find the actual last non-blank row to avoid issues with trailing empty rows
    last_data_row = ws.max_row
    while last_data_row > 0:
        row_data = [ws.cell(row=last_data_row, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None and str(v).strip() != "" for v in row_data):
            break
        last_data_row -= 1

    # Determine append position: one past the last data row (1-indexed, openpyxl 1-based)
    next_excel_row = last_data_row + 1

    # Group input rows by section name (preserve order)
    section_buckets: dict[str, list[dict]] = {}
    for r in rows:
        s = r.get("section") or "(default)"
        section_buckets.setdefault(s, []).append(r)

    for user_section, bucket in section_buckets.items():
        match = _find_section(user_section, sections)
        if match is None:
            # New section — write a section header row first
            new_section_idx = max(
                (int(s["last_tc_id"].strip().split("-")[0])
                 for s in sections
                 if s.get("last_tc_id") and "-" in s["last_tc_id"].strip()),
                default=0,
            ) + 1
            # Section header: Priority cell = numeric idx, Test Item cell = section name
            pri_idx = columns.get("Priority")
            item_idx = columns.get("Test Item")
            if pri_idx is not None:
                ws.cell(row=next_excel_row, column=pri_idx + 1, value=float(new_section_idx))
            if item_idx is not None:
                ws.cell(row=next_excel_row, column=item_idx + 1, value=user_section)
            next_excel_row += 1
            section_idx_str = str(new_section_idx)
            last_tc_id = None
        else:
            raw_last = match.get("last_tc_id") or ""
            last_tc_id = raw_last.strip() if raw_last else None
            section_idx_str = (
                last_tc_id.split("-")[0] if last_tc_id else "1"
            )

        for row_data in bucket:
            tc_id = _next_tc_id(section_idx_str, last_tc_id)
            last_tc_id = tc_id
            for col_name, col_idx in columns.items():
                if col_name == "TC_ID":
                    value = tc_id
                elif col_name == "Test Item" and "Test Item" not in row_data:
                    # Inherit from section name if not provided
                    value = user_section
                else:
                    value = row_data.get(col_name, "")
                cell = ws.cell(row=next_excel_row, column=col_idx + 1, value=value)
                cell.alignment = WRAP
            next_excel_row += 1

    # CRITICAL: save to new file, never to master
    wb.save(actual_output)
    return actual_output


def main() -> None:
    parser = argparse.ArgumentParser(description="Append TC rows to a master xlsx tab.")
    parser.add_argument("--master", type=Path, required=True)
    parser.add_argument("--tab", type=str, required=True)
    parser.add_argument("--rows", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    data = json.loads(args.rows.read_text(encoding="utf-8"))
    rows = data.get("rows", [])
    actual = append_to_master(args.master, args.tab, rows, args.output)
    print(f"Wrote {actual}")
    print(actual)


if __name__ == "__main__":
    main()
