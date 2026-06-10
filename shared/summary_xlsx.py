"""Write a summary workbook from a list of sheet specs.

Generic exporter for md+xlsx skills (qa-risk-analysis, qa-regression-scope).

CLI:
    python summary_xlsx.py --sheets sheets.json --output out.xlsx

sheets.json schema:
    {"sheets": [{"title": "리스크 매트릭스", "headers": ["영역", ...],
                 "rows": [["결제", ...], ...]}, ...]}

If output path already exists, appends "(2)", "(3)", ... suffix.
Prints the actual output path as the last stdout line.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from new_workbook import HEADER_FILL, HEADER_FONT, WRAP, _resolve_output_path  # noqa: E402

from openpyxl import Workbook  # noqa: E402


def write_summary_workbook(sheets: list[dict], output: Path) -> Path:
    """Create an xlsx with one tab per sheet spec. Returns actual path written."""
    if not sheets:
        raise ValueError("sheets must not be empty")
    actual_output = _resolve_output_path(output)

    wb = Workbook()
    wb.remove(wb.active)
    for spec in sheets:
        ws = wb.create_sheet(spec["title"])
        for col_idx, name in enumerate(spec.get("headers", []), start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = WRAP
        for row_idx, row in enumerate(spec.get("rows", []), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value).alignment = WRAP

    actual_output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(actual_output)
    return actual_output


def main() -> None:
    parser = argparse.ArgumentParser(description="Write a summary workbook.")
    parser.add_argument("--sheets", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    data = json.loads(args.sheets.read_text(encoding="utf-8"))
    actual = write_summary_workbook(data.get("sheets", []), args.output)
    print(f"Wrote {actual}")
    print(actual)


if __name__ == "__main__":
    main()
